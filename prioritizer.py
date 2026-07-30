"""
Automatic Lead Prioritization Engine
Applies scoring algorithms, ranks leads, adds sales-ready actions,
applies conditional formatting, and generates top-tier sales workbooks.
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import config


def has_valid_value(value):
    """
    Returns True only if a field contains a real, non-empty value.
    Filters out None, NaN, whitespace, and fallback strings like 'No Website', 'N/A', etc.
    """
    if value is None or pd.isna(value):
        return False

    val_str = str(value).strip().lower()
    if not val_str:
        return False

    invalid_patterns = {
        "none", "nan", "n/a", "null", "-",
        "no website", "no phone", "no email", "no whatsapp",
        "no facebook", "no instagram", "no linkedin"
    }

    if val_str in invalid_patterns:
        return False

    return True


def calculate_priority_score(row):
    score = 0

    # Email available +30
    email = row.get("Email")
    if has_valid_value(email) and "@" in str(email):
        score += 30

    # Website available +20
    website = row.get("Website")
    if has_valid_value(website):
        score += 20

    # Phone available +15
    phone = row.get("Phone")
    if has_valid_value(phone):
        score += 15

    # WhatsApp available +15
    whatsapp = row.get("WhatsApp")
    if has_valid_value(whatsapp):
        score += 15

    # LinkedIn available +8
    linkedin = row.get("LinkedIn")
    if has_valid_value(linkedin):
        score += 8

    # Facebook available +5
    facebook = row.get("Facebook")
    if has_valid_value(facebook):
        score += 5

    # Instagram available +5
    instagram = row.get("Instagram")
    if has_valid_value(instagram):
        score += 5

    # Rating ≥ 4.5 +8
    try:
        rating = float(row.get("Rating", 0) or 0)
        if rating >= 4.5:
            score += 8
    except Exception:
        pass

    # Reviews >100 (+10), >500 (+15), >1000 (+20)
    try:
        rev_str = str(row.get("Reviews", 0) or 0).replace(',', '')
        reviews = float(rev_str)
        if reviews > 1000:
            score += 20
        elif reviews > 500:
            score += 15
        elif reviews > 100:
            score += 10
    except Exception:
        pass

    # Working website +5 / Broken website +0
    web_status = row.get("Website Status")
    if has_valid_value(web_status) and str(web_status).strip().lower() == "working":
        score += 5

    return score


def get_priority_level(score):
    if score >= 95:
        return "A+"
    elif score >= 80:
        return "A"
    elif score >= 65:
        return "B"
    elif score >= 50:
        return "C"
    else:
        return "D"


def get_next_action(row):
    email = row.get("Email")
    whatsapp = row.get("WhatsApp")
    phone = row.get("Phone")
    website = row.get("Website")
    facebook = row.get("Facebook")
    instagram = row.get("Instagram")

    if has_valid_value(email) and "@" in str(email):
        return "Email First"
    elif has_valid_value(whatsapp):
        return "WhatsApp Outreach"
    elif has_valid_value(phone):
        return "Call Business"
    elif has_valid_value(website):
        return "Visit Website"
    elif has_valid_value(facebook) or has_valid_value(instagram):
        return "Social Media Outreach"
    else:
        return "Manual Research"


def apply_excel_formatting(filepath):
    """
    Applies conditional formatting and professional styling to a prioritized Excel workbook.
    """
    if not os.path.exists(filepath):
        return

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Style definitions
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Priority Level Fills
        level_styles = {
            "A+": {"fill": PatternFill(start_color="1E4620", end_color="1E4620", fill_type="solid"), "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF")},
            "A":  {"fill": PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid"), "font": Font(name="Calibri", size=11, bold=True, color="155724")},
            "B":  {"fill": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), "font": Font(name="Calibri", size=11, bold=True, color="856404")},
            "C":  {"fill": PatternFill(start_color="FFE8D6", end_color="FFE8D6", fill_type="solid"), "font": Font(name="Calibri", size=11, bold=True, color="853C00")},
            "D":  {"fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"), "font": Font(name="Calibri", size=11, bold=True, color="721C24")}
        }

        # Identify Column Indices
        headers = [cell.value for cell in ws[1]]
        level_col_idx = headers.index("Priority Level") + 1 if "Priority Level" in headers else None
        score_col_idx = headers.index("Priority Score") + 1 if "Priority Score" in headers else None
        rank_col_idx = headers.index("Rank") + 1 if "Rank" in headers else None
        action_col_idx = headers.index("Next Action") + 1 if "Next Action" in headers else None

        # Format Headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Format Rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = left_align

            # Highlight Priority Level Column
            if level_col_idx:
                l_cell = ws.cell(row=row_idx, column=level_col_idx)
                val = str(l_cell.value or "").strip()
                if val in level_styles:
                    l_cell.fill = level_styles[val]["fill"]
                    l_cell.font = level_styles[val]["font"]
                    l_cell.alignment = center_align

            if rank_col_idx:
                ws.cell(row=row_idx, column=rank_col_idx).alignment = center_align
            if score_col_idx:
                ws.cell(row=row_idx, column=score_col_idx).alignment = center_align
            if action_col_idx:
                ws.cell(row=row_idx, column=action_col_idx).alignment = center_align

        wb.save(filepath)
    except Exception as e:
        print(f"Excel formatting notice for {filepath}: {e}", flush=True)


def update_summary_workbook(df):
    """
    Appends a 'Priority Analytics' sheet to outputs/summary.xlsx without modifying Sheet 1.
    """
    summary_path = config.EXCEL_SUMMARY
    if not os.path.exists(summary_path):
        return

    try:
        wb = openpyxl.load_workbook(summary_path)
        sheet_name = "Priority Analytics"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        # Style Definitions
        title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_header_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")

        # 1. KPI Overview Table
        ws.cell(row=1, column=1, value="PRIORITY ANALYTICS OVERVIEW").font = title_font

        total_biz = len(df)
        avg_score = round(df["Priority Score"].mean(), 1) if not df.empty else 0
        max_score = df["Priority Score"].max() if not df.empty else 0
        min_score = df["Priority Score"].min() if not df.empty else 0

        # Safe Rating Avg
        try:
            avg_rating = round(pd.to_numeric(df["Rating"], errors='coerce').mean(), 2)
        except Exception:
            avg_rating = 0.0

        kpis = [
            ("Total Businesses", total_biz),
            ("Average Priority Score", avg_score),
            ("Highest Score", max_score),
            ("Lowest Score", min_score),
            ("Average Rating", avg_rating if not pd.isna(avg_rating) else 0.0),
            ("Businesses with Email", sum(1 for e in df["Email"] if has_valid_value(e) and "@" in str(e))),
            ("Businesses with Website", sum(1 for w in df["Website"] if has_valid_value(w))),
            ("Businesses with Phone", sum(1 for p in df["Phone"] if has_valid_value(p))),
            ("Businesses with WhatsApp", sum(1 for w in df["WhatsApp"] if has_valid_value(w))),
            ("Businesses with LinkedIn", sum(1 for l in df["LinkedIn"] if has_valid_value(l))),
            ("Businesses with Facebook", sum(1 for f in df["Facebook"] if has_valid_value(f))),
            ("Businesses with Instagram", sum(1 for i in df["Instagram"] if has_valid_value(i)))
        ]

        ws.cell(row=3, column=1, value="Metric").fill = header_fill
        ws.cell(row=3, column=1).font = header_font
        ws.cell(row=3, column=2, value="Value").fill = header_fill
        ws.cell(row=3, column=2).font = header_font

        for r_idx, (m, v) in enumerate(kpis, start=4):
            ws.cell(row=r_idx, column=1, value=m)
            ws.cell(row=r_idx, column=2, value=v)

        # 2. Priority Distribution Table
        ws.cell(row=17, column=1, value="PRIORITY DISTRIBUTION").font = sub_header_font
        ws.cell(row=18, column=1, value="Priority Level").fill = header_fill
        ws.cell(row=18, column=1).font = header_font
        ws.cell(row=18, column=2, value="Count").fill = header_fill
        ws.cell(row=18, column=2).font = header_font

        dist = df["Priority Level"].value_counts().to_dict()
        for r_idx, lvl in enumerate(["A+", "A", "B", "C", "D"], start=19):
            ws.cell(row=r_idx, column=1, value=lvl)
            ws.cell(row=r_idx, column=2, value=dist.get(lvl, 0))

        # 3. Top 10 Businesses Table
        ws.cell(row=26, column=1, value="TOP 10 RANKED BUSINESSES").font = sub_header_font
        t10_cols = ["Rank", "Business Name", "Source Category", "Search Location", "Priority Score", "Priority Level", "Next Action"]
        for c_idx, col in enumerate(t10_cols, start=1):
            cell = ws.cell(row=27, column=c_idx, value=col)
            cell.fill = header_fill
            cell.font = header_font

        top_10_df = df.head(10)
        for r_idx, (_, row) in enumerate(top_10_df.iterrows(), start=28):
            for c_idx, col in enumerate(t10_cols, start=1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(col, ""))

        wb.save(summary_path)
    except Exception as e:
        print(f"Summary analytics notice: {e}", flush=True)


def process_lead_prioritization():
    """
    Main entry point: Reads all_leads.xlsx, calculates priority scores,
    ranks businesses, exports prioritized and top-N workbooks, and updates summary.
    Executes safely as an independent post-processing step.
    """
    input_file = config.EXCEL_ALL
    if not os.path.exists(input_file):
        print(f"Prioritization Notice: {input_file} does not exist yet.", flush=True)
        return

    try:
        df = pd.read_excel(input_file)
        if df.empty:
            print("Prioritization Notice: all_leads.xlsx is empty.", flush=True)
            return

        print("\n========================================", flush=True)
        print("STARTING AUTOMATIC LEAD PRIORITIZATION", flush=True)
        print("========================================", flush=True)

        # 1. Calculate Priority Score, Priority Level, Next Action
        df["Priority Score"] = df.apply(calculate_priority_score, axis=1)
        df["Priority Level"] = df["Priority Score"].apply(get_priority_level)
        df["Next Action"] = df.apply(get_next_action, axis=1)

        # Helper numeric columns for sorting
        df["_rating_num"] = pd.to_numeric(df["Rating"], errors='coerce').fillna(0.0)
        df["_reviews_num"] = pd.to_numeric(df["Reviews"].astype(str).str.replace(',', ''), errors='coerce').fillna(0.0)

        # 2. Multi-column Sorting: Score (Desc), Rating (Desc), Reviews (Desc), Name (Asc)
        df = df.sort_values(
            by=["Priority Score", "_rating_num", "_reviews_num", "Business Name"],
            ascending=[False, False, False, True]
        ).reset_index(drop=True)

        # 3. Assign 1-based Rank
        df["Rank"] = range(1, len(df) + 1)

        # Drop temporary helper numeric columns
        df = df.drop(columns=["_rating_num", "_reviews_num"])

        # Re-order columns so new sales columns append at the end
        base_cols = [c for c in df.columns if c not in ["Priority Score", "Priority Level", "Rank", "Next Action"]]
        final_cols = base_cols + ["Priority Score", "Priority Level", "Rank", "Next Action"]
        df_final = df[final_cols]

        # 4. Generate Output Workbooks
        out_prioritized = os.path.join(config.OUTPUTS_DIR, "all_leads_prioritized.xlsx")
        out_top_25 = os.path.join(config.OUTPUTS_DIR, "top_25_leads.xlsx")
        out_top_50 = os.path.join(config.OUTPUTS_DIR, "top_50_leads.xlsx")
        out_top_100 = os.path.join(config.OUTPUTS_DIR, "top_100_leads.xlsx")

        # Atomic writes to prevent workbook corruption
        def atomic_export(dataframe, path):
            tmp_p = path + ".tmp.xlsx"
            dataframe.to_excel(tmp_p, index=False)
            os.replace(tmp_p, path)
            apply_excel_formatting(path)

        atomic_export(df_final, out_prioritized)
        atomic_export(df_final.head(25), out_top_25)
        atomic_export(df_final.head(50), out_top_50)
        atomic_export(df_final.head(100), out_top_100)

        # 5. Update summary.xlsx with Priority Analytics sheet
        update_summary_workbook(df_final)

        print(f"Prioritized Leads Saved : {out_prioritized}", flush=True)
        print(f"Top 25 Leads Exported   : {out_top_25}", flush=True)
        print(f"Top 50 Leads Exported   : {out_top_50}", flush=True)
        print(f"Top 100 Leads Exported  : {out_top_100}", flush=True)
        print("Priority Analytics added to summary.xlsx", flush=True)
        print("========================================\n", flush=True)

    except Exception as e:
        print(f"Prioritization Warning: {e}", flush=True)


if __name__ == "__main__":
    process_lead_prioritization()
