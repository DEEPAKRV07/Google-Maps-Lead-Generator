"""
Dual Dashboard Generator Module (dashboard.py)
Queries database/db.sqlite3 to generate:
1. outputs/Dashboard.xlsx (Excel KPI Workbook)
2. outputs/dashboard.html (Standalone Web Browser Dashboard)
"""

import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config
import logger
import database.db as db

HTML_DASHBOARD_FILE = os.path.join(config.OUTPUTS_DIR, "dashboard.html")
EXCEL_DASHBOARD_FILE = os.path.join(config.OUTPUTS_DIR, "Dashboard.xlsx")


def get_dashboard_metrics():
    """
    Queries database/db.sqlite3 and returns aggregated KPI metrics dictionary.
    """
    conn = db.get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM businesses")
    total_leads = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE phone IS NOT NULL AND phone != ''")
    has_phone = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE website IS NOT NULL AND website != ''")
    has_website = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE email IS NOT NULL AND email != ''")
    has_email = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE whatsapp IS NOT NULL AND whatsapp != ''")
    has_whatsapp = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE facebook IS NOT NULL AND facebook != ''")
    has_facebook = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE instagram IS NOT NULL AND instagram != ''")
    has_instagram = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM businesses WHERE linkedin IS NOT NULL AND linkedin != ''")
    has_linkedin = cursor.fetchone()[0]

    cursor.execute("SELECT priority_level, COUNT(*) FROM businesses GROUP BY priority_level")
    level_counts = dict(cursor.fetchall())

    cursor.execute("SELECT next_action, COUNT(*) FROM businesses GROUP BY next_action ORDER BY COUNT(*) DESC")
    action_counts = dict(cursor.fetchall())

    cursor.execute("SELECT source_category, COUNT(*) FROM businesses GROUP BY source_category ORDER BY COUNT(*) DESC")
    category_counts = dict(cursor.fetchall())

    conn.close()

    return {
        "total_leads": total_leads,
        "has_phone": has_phone,
        "has_website": has_website,
        "has_email": has_email,
        "has_whatsapp": has_whatsapp,
        "has_facebook": has_facebook,
        "has_instagram": has_instagram,
        "has_linkedin": has_linkedin,
        "level_counts": level_counts,
        "action_counts": action_counts,
        "category_counts": category_counts
    }


def generate_html_dashboard(metrics):
    """
    Generates outputs/dashboard.html standalone HTML web dashboard.
    """
    m = metrics
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Lead Generator v2.0 Enterprise Dashboard</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-primary: #0F172A;
            --bg-card: #1E293B;
            --text-primary: #F8FAFC;
            --text-secondary: #94A3B8;
            --accent-green: #10B981;
            --accent-blue: #3B82F6;
            --accent-purple: #8B5CF6;
            --accent-amber: #F59E0B;
            --border-color: #334155;
        }}
        body {{
            font-family: 'Inter', sans-serif;
            background-color: var(--bg-primary);
            color: var(--text-primary);
            margin: 0;
            padding: 24px;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 32px;
            padding-bottom: 16px;
            border-bottom: 1px solid var(--border-color);
        }}
        .header h1 {{
            font-size: 24px;
            font-weight: 700;
            margin: 0;
            background: linear-gradient(135deg, #3B82F6, #10B981);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        .badge {{
            background: rgba(16, 185, 129, 0.2);
            color: var(--accent-green);
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 600;
        }}
        .grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 20px;
            margin-bottom: 32px;
        }}
        .card {{
            background-color: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }}
        .card .title {{
            font-size: 13px;
            color: var(--text-secondary);
            margin-bottom: 8px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .card .value {{
            font-size: 28px;
            font-weight: 700;
            color: var(--text-primary);
        }}
        .card .subtitle {{
            font-size: 12px;
            color: var(--accent-green);
            margin-top: 4px;
        }}
        .section-title {{
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 16px;
            color: var(--text-primary);
        }}
        .table-container {{
            background-color: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            overflow: hidden;
            margin-bottom: 32px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
        }}
        th, td {{
            padding: 14px 20px;
            border-bottom: 1px solid var(--border-color);
        }}
        th {{
            background-color: rgba(15, 23, 42, 0.6);
            color: var(--text-secondary);
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        tr:last-child td {{
            border-bottom: none;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div>
            <h1>🚀 Lead Generator v2.0 Enterprise Dashboard</h1>
            <div style="color: var(--text-secondary); font-size: 14px; margin-top: 4px;">Real-Time Analytics & Master Lead Intelligence</div>
        </div>
        <span class="badge">v2.0.0 Live Sync</span>
    </div>

    <div class="grid">
        <div class="card">
            <div class="title">Total Businesses</div>
            <div class="value">{m['total_leads']}</div>
            <div class="subtitle">SQLite Master DB</div>
        </div>
        <div class="card">
            <div class="title">Phone Numbers</div>
            <div class="value">{m['has_phone']}</div>
            <div class="subtitle">{round(m['has_phone']/max(m['total_leads'],1)*100, 1)}% Coverage</div>
        </div>
        <div class="card">
            <div class="title">Websites</div>
            <div class="value">{m['has_website']}</div>
            <div class="subtitle">{round(m['has_website']/max(m['total_leads'],1)*100, 1)}% Coverage</div>
        </div>
        <div class="card">
            <div class="title">Validated Emails</div>
            <div class="value">{m['has_email']}</div>
            <div class="subtitle">{round(m['has_email']/max(m['total_leads'],1)*100, 1)}% Coverage</div>
        </div>
    </div>

    <div class="section-title">📊 Sales Strategy & Next Action Breakdown</div>
    <div class="table-container">
        <table>
            <thead>
                <tr>
                    <th>Next Action Strategy</th>
                    <th>Lead Count</th>
                    <th>Share of Total</th>
                </tr>
            </thead>
            <tbody>
"""
    for act, count in m['action_counts'].items():
        pct = round(count / max(m['total_leads'], 1) * 100, 1)
        html_content += f"""
                <tr>
                    <td><strong>{act}</strong></td>
                    <td>{count}</td>
                    <td>{pct}%</td>
                </tr>
"""
    html_content += """
            </tbody>
        </table>
    </div>

    <div class="section-title">🎯 Priority Level Distribution</div>
    <div class="table-container">
        <table>
            <thead>
                <tr>
                    <th>Priority Level</th>
                    <th>Description</th>
                    <th>Lead Count</th>
                </tr>
            </thead>
            <tbody>
"""
    levels = [
        ("A+", "Top Target (Score >= 50, High Contactability)", m['level_counts'].get('A+', 0)),
        ("A", "High Priority Target (Score 35-49)", m['level_counts'].get('A', 0)),
        ("B", "Medium Priority Target (Score 20-34)", m['level_counts'].get('B', 0)),
        ("C", "Low Priority Target (Score 10-19)", m['level_counts'].get('C', 0)),
        ("D", "Minimal Contact Info (Score < 10)", m['level_counts'].get('D', 0))
    ]
    for lvl, desc, cnt in levels:
        html_content += f"""
                <tr>
                    <td><strong>{lvl}</strong></td>
                    <td>{desc}</td>
                    <td>{cnt}</td>
                </tr>
"""
    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""
    try:
        with open(HTML_DASHBOARD_FILE, "w", encoding="utf-8") as f:
            f.write(html_content)
        logger.info("database", f"Generated standalone HTML dashboard: {HTML_DASHBOARD_FILE}")
    except Exception as e:
        logger.error("database", f"Generate HTML dashboard error: {e}")


def generate_excel_dashboard(metrics):
    """
    Generates outputs/Dashboard.xlsx Excel workbook using openpyxl.
    """
    m = metrics
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Overview"

    # Styling
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    ws.cell(row=1, column=1, value="Google Maps Lead Generator v2.0 Dashboard").font = font_title

    headers = ["Metric", "Value", "Percentage"]
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_header

    rows = [
        ("Total Businesses", m['total_leads'], "100.0%"),
        ("Businesses with Phone", m['has_phone'], f"{round(m['has_phone']/max(m['total_leads'],1)*100, 1)}%"),
        ("Businesses with Website", m['has_website'], f"{round(m['has_website']/max(m['total_leads'],1)*100, 1)}%"),
        ("Businesses with Email", m['has_email'], f"{round(m['has_email']/max(m['total_leads'],1)*100, 1)}%"),
        ("Businesses with WhatsApp", m['has_whatsapp'], f"{round(m['has_whatsapp']/max(m['total_leads'],1)*100, 1)}%"),
        ("Businesses with Instagram", m['has_instagram'], f"{round(m['has_instagram']/max(m['total_leads'],1)*100, 1)}%"),
        ("Businesses with Facebook", m['has_facebook'], f"{round(m['has_facebook']/max(m['total_leads'],1)*100, 1)}%"),
        ("Businesses with LinkedIn", m['has_linkedin'], f"{round(m['has_linkedin']/max(m['total_leads'],1)*100, 1)}%")
    ]

    for row_idx, r_data in enumerate(rows, 4):
        for col_idx, val in enumerate(r_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    try:
        wb.save(EXCEL_DASHBOARD_FILE)
        logger.info("database", f"Generated Excel dashboard: {EXCEL_DASHBOARD_FILE}")
    except Exception as e:
        logger.error("database", f"Generate Excel dashboard error: {e}")


def generate_all_dashboards():
    """
    Main runner to generate both HTML and Excel dashboards.
    """
    metrics = get_dashboard_metrics()
    generate_html_dashboard(metrics)
    generate_excel_dashboard(metrics)
